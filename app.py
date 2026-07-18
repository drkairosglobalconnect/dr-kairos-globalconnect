import sqlite3
import razorpay
import os
from openpyxl import Workbook

from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, send_file
from flask_mail import Mail, Message
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
from werkzeug.utils import secure_filename

app = Flask(__name__)

app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True

app.config["MAIL_USERNAME"] = "drkairosglobalconnect@gmail.com"
app.config["MAIL_PASSWORD"] = "sdvywoanilmpmgw"

mail = Mail(app)

client = razorpay.Client(auth=(
    "rzp_test_TEwktCUYs0SlOF",
    "odgRe3XKlIeNcxnyB4s10oi0"
))

from flask import send_from_directory
import os
from werkzeug.utils import secure_filename

from flask import Flask, render_template, request, redirect, url_for, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_


UPLOAD_FOLDER = "uploads"

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

app.secret_key = "drkairos_secret_key"

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///drkairos.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

class Doctor(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    full_name = db.Column(db.String(100))
    email = db.Column(db.String(100))
    mobile = db.Column(db.String(20))
    specialty = db.Column(db.String(100))
    hospital = db.Column(db.String(100))
    experience = db.Column(db.Integer)

    password = db.Column(db.String(200))

    degree_certificate = db.Column(db.String(200))
    registration_certificate = db.Column(db.String(200))

class Student(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    full_name = db.Column(db.String(100))

    dob = db.Column(db.String(20))

    gender = db.Column(db.String(20))

    college = db.Column(db.String(100))

    university = db.Column(db.String(100))

    course = db.Column(db.String(100))

    year = db.Column(db.String(20))

    email = db.Column(db.String(100))

    mobile = db.Column(db.String(20))

    password = db.Column(db.String(200))

    whatsapp = db.Column(db.String(20))

class Hospital(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    hospital_name = db.Column(db.String(150))
    hospital_type = db.Column(db.String(100))
    beds = db.Column(db.Integer)
    doctors = db.Column(db.Integer)
    specialties = db.Column(db.String(300))

    contact_name = db.Column(db.String(100))
    designation = db.Column(db.String(100))
    email = db.Column(db.String(100))
    mobile = db.Column(db.String(20))

    password = db.Column(db.String(200))

    city = db.Column(db.String(100))
    state = db.Column(db.String(100))
    country = db.Column(db.String(100))

class Pharma(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    company_name = db.Column(db.String(150))
    registration_number = db.Column(db.String(100))
    website = db.Column(db.String(150))

    contact_name = db.Column(db.String(100))
    designation = db.Column(db.String(100))
    email = db.Column(db.String(100))
    mobile = db.Column(db.String(20))

    password = db.Column(db.String(200))

    category = db.Column(db.String(100))

class Organizer(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    organization_name = db.Column(db.String(150))
    website = db.Column(db.String(150))
    contact_person = db.Column(db.String(100))
    email = db.Column(db.String(100))
    mobile = db.Column(db.String(20))

    password = db.Column(db.String(200))

    conference_name = db.Column(db.String(150))
    specialty = db.Column(db.String(100))
    country = db.Column(db.String(100))
    venue = db.Column(db.String(150))
    conference_date = db.Column(db.String(50))

class Conference(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    conference_name = db.Column(db.String(200))
    specialty = db.Column(db.String(100))
    country = db.Column(db.String(100))
    venue = db.Column(db.String(200))
    conference_date = db.Column(db.String(50))
    organizer = db.Column(db.String(150))
    registration_fee = db.Column(db.String(50))

class ConferenceRegistration(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    doctor_id = db.Column(db.Integer)

    conference_id = db.Column(db.Integer)

    payment_status = db.Column(db.String(50), default="Pending")

    payment_id = db.Column(db.String(150))


@app.route("/")
def home():
    return render_template("index.html")

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/services")
def services():
    return render_template("services.html")

@app.route("/conferences")
def conferences():
    return render_template("conferences.html")

@app.route("/registration")
def registration():
    return render_template("registration.html")

@app.route("/doctor", methods=["GET", "POST"])
def doctor():

    if request.method == "POST":

        # Upload files
        degree = request.files["degree"]
        registration = request.files["registration_certificate"]

        degree_file = ""
        registration_file = ""

        if degree and degree.filename:
            degree_file = secure_filename(degree.filename)
            degree.save(os.path.join(app.config["UPLOAD_FOLDER"], degree_file))

        if registration and registration.filename:
            registration_file = secure_filename(registration.filename)
            registration.save(os.path.join(app.config["UPLOAD_FOLDER"], registration_file))

        # Create Doctor object
        doctor = Doctor(
            full_name=request.form["full_name"],
            email=request.form["email"],
            mobile=request.form["mobile"],
            specialty=request.form["specialty"],
            hospital=request.form["hospital"],
            experience=request.form["experience"],
            password=request.form["password"],
            degree_certificate=degree_file,
            registration_certificate=registration_file
        )

        # Save to database
        db.session.add(doctor)
        db.session.commit()

        # Send email
        msg = Message(
            "New Doctor Registration - Dr Kairos GlobalConnect",
            sender=os.environ.get("MAIL_USERNAME"),
            recipients=[os.environ.get("MAIL_USERNAME")]
        )

        msg.body = f"""
New Doctor Registration Received

Name: {doctor.full_name}
Email: {doctor.email}
Mobile: {doctor.mobile}
Specialization: {doctor.specialty}
Hospital: {doctor.hospital}
Experience: {doctor.experience} years
"""

        try:
            mail.send(msg)
        except Exception as e:
            print(e)

        return redirect(url_for("doctor"))

    return render_template("doctor.html")

@app.route("/student", methods=["GET", "POST"])
def student():

    if request.method == "POST":

        student = Student(

    full_name=request.form["full_name"],
    dob=request.form["dob"],
    gender=request.form["gender"],
    college=request.form["college"],
    university=request.form["university"],
    course=request.form["course"],
    year=request.form["year"],
    email=request.form["email"],
    mobile=request.form["mobile"],
    whatsapp=request.form["whatsapp"],

    password=request.form["password"]

)

        db.session.add(student)
        db.session.commit()

        return redirect(url_for("student"))

    return render_template("student.html")

@app.route("/hospital", methods=["GET", "POST"])
def hospital():

    if request.method == "POST":

        hospital = Hospital(

            hospital_name=request.form["hospital_name"],
            hospital_type=request.form["hospital_type"],
            beds=request.form["beds"],
            doctors=request.form["doctors"],
            specialties=request.form["specialties"],

            contact_name=request.form["contact_name"],
            designation=request.form["designation"],
            email=request.form["email"],
            mobile=request.form["mobile"],

            password=request.form["password"],
   
            city=request.form["city"],
            state=request.form["state"],
            country=request.form["country"]

        )

        db.session.add(hospital)
        db.session.commit()

        return redirect(url_for("hospital"))

    return render_template("hospital.html")

@app.route("/pharma", methods=["GET", "POST"])
def pharma():

    if request.method == "POST":

        pharma = Pharma(
            company_name=request.form["company_name"],
            registration_number=request.form["registration_number"],
            website=request.form["website"],
            contact_name=request.form["contact_name"],
            designation=request.form["designation"],
            email=request.form["email"],
            mobile=request.form["mobile"],

            password=request.form["password"],

            category=request.form["category"]
        )

        db.session.add(pharma)
        db.session.commit()

        return redirect(url_for("pharma"))

    return render_template("pharma.html")

@app.route("/organizer", methods=["GET", "POST"])
def organizer():

    if request.method == "POST":

        organizer = Organizer(

            organization_name=request.form["organization_name"],
            website=request.form["website"],
            contact_person=request.form["contact_person"],
            email=request.form["email"],
            mobile=request.form["mobile"],

            password=request.form["password"],

            conference_name=request.form["conference_name"],
            specialty=request.form["specialty"],
            country=request.form["country"],
            venue=request.form["venue"],
            conference_date=request.form["conference_date"]

        )

        db.session.add(organizer)
        db.session.commit()

        return redirect(url_for("organizer"))

    return render_template("organizer.html")

@app.route("/add_conference", methods=["GET", "POST"])
def add_conference():

    if "admin" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":

        conference = Conference(

            conference_name=request.form["conference_name"],
            specialty=request.form["specialty"],
            country=request.form["country"],
            venue=request.form["venue"],
            conference_date=request.form["conference_date"],
            organizer=request.form["organizer"],
            registration_fee=request.form["registration_fee"]

        )

        db.session.add(conference)
        db.session.commit()

        return redirect(url_for("conference_list"))

    return render_template("add_conference.html")

@app.route("/conference_list")
def conference_list():

    conferences = Conference.query.all()

    return render_template(
        "conference_list.html",
        conferences=conferences
    )

@app.route("/payment/<int:id>")
def payment(id):

    conference = Conference.query.get_or_404(id)

    amount = int(float(conference.registration_fee) * 100)

    order = client.order.create({
        "amount": amount,
        "currency": "INR",
        "payment_capture": 1
    })

    return render_template(
        "payment.html",
        conference=conference,
        order=order,
        key_id="rzp_test_TEwO4Vga9LJ7b2"
    )

@app.route("/payment_success", methods=["POST"])
def payment_success():

    if "doctor" not in session:
        return redirect(url_for("doctor_login"))

    payment_id = request.form["razorpay_payment_id"]
    order_id = request.form["razorpay_order_id"]
    signature = request.form["razorpay_signature"]
    conference_id = request.form["conference_id"]

    params = {
        "razorpay_order_id": order_id,
        "razorpay_payment_id": payment_id,
        "razorpay_signature": signature
    }

    try:
        client.utility.verify_payment_signature(params)

    except razorpay.errors.SignatureVerificationError:
        return "Payment Verification Failed!"

    registration = ConferenceRegistration(
        doctor_id=session["doctor"],
        conference_id=conference_id,
        payment_status="Paid",
        payment_id=payment_id
    )

    db.session.add(registration)
    db.session.commit()

    return redirect(url_for("my_conferences"))


@app.route("/edit_conference/<int:id>", methods=["GET", "POST"])
def edit_conference(id):

    conference = Conference.query.get_or_404(id)

    if request.method == "POST":

        conference.conference_name = request.form["conference_name"]
        conference.specialty = request.form["specialty"]
        conference.country = request.form["country"]
        conference.venue = request.form["venue"]
        conference.conference_date = request.form["conference_date"]
        conference.organizer = request.form["organizer"]
        conference.registration_fee = request.form["registration_fee"]

        db.session.commit()

        return redirect(url_for("conference_list"))

    return render_template(
        "edit_conference.html",
        conference=conference
    )

@app.route("/delete_conference/<int:id>")
def delete_conference(id):

    conference = Conference.query.get_or_404(id)

    db.session.delete(conference)
    db.session.commit()

    return redirect(url_for("conference_list"))

@app.route("/partners")
def partners():
    return render_template("partners.html")

@app.route("/founder")
def founder():
    return render_template("founder.html")

@app.route("/contact")
def contact():
    return render_template("contact.html")

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        if username == "admin" and password == "123456":

            session["admin"] = True
            return redirect(url_for("admin"))

        return "Invalid Username or Password"

    return render_template("login.html")

@app.route("/admin")
def admin():

    if "admin" not in session:
        return redirect(url_for("login"))

    search = request.args.get("search", "")

    if search:
        doctors = Doctor.query.filter(
            or_(
                Doctor.full_name.ilike(f"%{search}%"),
                Doctor.email.ilike(f"%{search}%"),
                Doctor.specialty.ilike(f"%{search}%"),
                Doctor.hospital.ilike(f"%{search}%")
            )
        ).all()
    else:
        doctors = Doctor.query.all()

    students = Student.query.all()
    hospitals = Hospital.query.all()
    pharmas = Pharma.query.all()
    organizers = Organizer.query.all()

    return render_template(
        "admin.html",
        doctors=doctors,
        students=students,
        hospitals=hospitals,
        pharmas=pharmas,
        organizers=organizers,
        search=search
    )

@app.route("/export_doctors")
def export_doctors():

    if "admin" not in session:
        return redirect(url_for("login"))

    doctors = Doctor.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Doctors"

    ws.append([
        "ID",
        "Name",
        "Email",
        "Mobile",
        "Specialty",
        "Hospital",
        "Experience"
    ])

    for doctor in doctors:
        ws.append([
            doctor.id,
            doctor.full_name,
            doctor.email,
            doctor.mobile,
            doctor.specialty,
            doctor.hospital,
            doctor.experience
        ])

    file_path = "doctors.xlsx"
    wb.save(file_path)

    return send_file(
        file_path,
        as_attachment=True
    )

@app.route("/logout")
def logout():

    session.pop("admin", None)

    return redirect(url_for("login"))

@app.route("/edit_doctor/<int:id>", methods=["GET", "POST"])
def edit_doctor(id):

    doctor = Doctor.query.get_or_404(id)

    if request.method == "POST":

        doctor.full_name = request.form["full_name"]
        doctor.email = request.form["email"]
        doctor.mobile = request.form["mobile"]
        doctor.specialty = request.form["specialty"]
        doctor.hospital = request.form["hospital"]
        doctor.experience = request.form["experience"]

        db.session.commit()

        return redirect(url_for("admin"))

    return render_template("edit_doctor.html", doctor=doctor)

@app.route("/delete_doctor/<int:id>")
def delete_doctor(id):

    doctor = Doctor.query.get_or_404(id)

    db.session.delete(doctor)
    db.session.commit()

    return redirect(url_for("admin"))
@app.route("/edit_student/<int:id>", methods=["GET", "POST"])
def edit_student(id):

    student = Student.query.get_or_404(id)

    if request.method == "POST":

        student.full_name = request.form["full_name"]
        student.dob = request.form["dob"]
        student.gender = request.form["gender"]
        student.college = request.form["college"]
        student.university = request.form["university"]
        student.course = request.form["course"]
        student.year = request.form["year"]
        student.email = request.form["email"]
        student.mobile = request.form["mobile"]
        student.whatsapp = request.form["whatsapp"]

        db.session.commit()

        return redirect(url_for("admin"))

    return render_template("edit_student.html", student=student)


@app.route("/delete_student/<int:id>")
def delete_student(id):

    student = Student.query.get_or_404(id)

    db.session.delete(student)
    db.session.commit()

    return redirect(url_for("admin"))

@app.route("/export_students")
def export_students():

    if "admin" not in session:
        return redirect(url_for("login"))

    students = Student.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append([
        "ID",
        "Name",
        "DOB",
        "Gender",
        "College",
        "University",
        "Course",
        "Year",
        "Email",
        "Mobile"
    ])

    for student in students:
        ws.append([
            student.id,
            student.full_name,
            student.dob,
            student.gender,
            student.college,
            student.university,
            student.course,
            student.year,
            student.email,
            student.mobile
        ])

    file_path = "students.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

@app.route("/edit_hospital/<int:id>", methods=["GET", "POST"])
def edit_hospital(id):

    hospital = Hospital.query.get_or_404(id)

    if request.method == "POST":

        hospital.hospital_name = request.form["hospital_name"]
        hospital.hospital_type = request.form["hospital_type"]
        hospital.beds = request.form["beds"]
        hospital.doctors = request.form["doctors"]
        hospital.specialties = request.form["specialties"]

        hospital.contact_name = request.form["contact_name"]
        hospital.designation = request.form["designation"]
        hospital.email = request.form["email"]
        hospital.mobile = request.form["mobile"]

        hospital.city = request.form["city"]
        hospital.state = request.form["state"]
        hospital.country = request.form["country"]

        db.session.commit()

        return redirect(url_for("admin"))

    return render_template("edit_hospital.html", hospital=hospital)

@app.route("/delete_hospital/<int:id>")
def delete_hospital(id):

    hospital = Hospital.query.get_or_404(id)

    db.session.delete(hospital)
    db.session.commit()

    return redirect(url_for("admin"))

@app.route("/export_hospitals")
def export_hospitals():

    if "admin" not in session:
        return redirect(url_for("login"))

    hospitals = Hospital.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hospitals"

    ws.append([
        "ID",
        "Hospital Name",
        "Type",
        "Beds",
        "Doctors",
        "Specialties",
        "Contact",
        "Email",
        "Mobile",
        "City",
        "State",
        "Country"
    ])

    for hospital in hospitals:
        ws.append([
            hospital.id,
            hospital.hospital_name,
            hospital.hospital_type,
            hospital.beds,
            hospital.doctors,
            hospital.specialties,
            hospital.contact_name,
            hospital.email,
            hospital.mobile,
            hospital.city,
            hospital.state,
            hospital.country
        ])

    file_path = "hospitals.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

@app.route("/edit_pharma/<int:id>", methods=["GET", "POST"])
def edit_pharma(id):

    pharma = Pharma.query.get_or_404(id)

    if request.method == "POST":

        pharma.company_name = request.form["company_name"]
        pharma.registration_number = request.form["registration_number"]
        pharma.website = request.form["website"]

        pharma.contact_name = request.form["contact_name"]
        pharma.designation = request.form["designation"]
        pharma.email = request.form["email"]
        pharma.mobile = request.form["mobile"]

        pharma.category = request.form["category"]

        db.session.commit()

        return redirect(url_for("admin"))

    return render_template("edit_pharma.html", pharma=pharma)

@app.route("/delete_pharma/<int:id>")
def delete_pharma(id):

    pharma = Pharma.query.get_or_404(id)

    db.session.delete(pharma)
    db.session.commit()

    return redirect(url_for("admin"))

@app.route("/export_pharma")
def export_pharma():

    if "admin" not in session:
        return redirect(url_for("login"))

    pharmas = Pharma.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pharma"

    ws.append([
        "ID",
        "Company Name",
        "Registration Number",
        "Website",
        "Contact Name",
        "Email",
        "Mobile",
        "Category"
    ])

    for pharma in pharmas:
        ws.append([
            pharma.id,
            pharma.company_name,
            pharma.registration_number,
            pharma.website,
            pharma.contact_name,
            pharma.email,
            pharma.mobile,
            pharma.category
        ])

    file_path = "pharma.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

@app.route("/edit_organizer/<int:id>", methods=["GET", "POST"])
def edit_organizer(id):

    organizer = Organizer.query.get_or_404(id)

    if request.method == "POST":

        organizer.organization_name = request.form["organization_name"]
        organizer.website = request.form["website"]
        organizer.contact_person = request.form["contact_person"]
        organizer.email = request.form["email"]
        organizer.mobile = request.form["mobile"]

        organizer.conference_name = request.form["conference_name"]
        organizer.specialty = request.form["specialty"]
        organizer.country = request.form["country"]
        organizer.venue = request.form["venue"]
        organizer.conference_date = request.form["conference_date"]

        db.session.commit()

        return redirect(url_for("admin"))

    return render_template("edit_organizer.html", organizer=organizer)

@app.route("/delete_organizer/<int:id>")
def delete_organizer(id):

    organizer = Organizer.query.get_or_404(id)

    db.session.delete(organizer)
    db.session.commit()

    return redirect(url_for("admin"))

@app.route("/export_organizers")
def export_organizers():

    if "admin" not in session:
        return redirect(url_for("login"))

    organizers = Organizer.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Organizers"

    ws.append([
        "ID",
        "Organization Name",
        "Website",
        "Contact Person",
        "Email",
        "Mobile",
        "Conference Name",
        "Specialty",
        "Country",
        "Venue",
        "Conference Date"
    ])

    for organizer in organizers:
        ws.append([
            organizer.id,
            organizer.organization_name,
            organizer.website,
            organizer.contact_person,
            organizer.email,
            organizer.mobile,
            organizer.conference_name,
            organizer.specialty,
            organizer.country,
            organizer.venue,
            organizer.conference_date
        ])

    file_path = "organizers.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

@app.route("/doctor_login", methods=["GET", "POST"])
def doctor_login():

    if request.method == "POST":

        email = request.form["email"]
        password = request.form["password"]

        doctor = Doctor.query.filter_by(
            email=email,
            password=password
        ).first()

        if doctor:

            session["doctor"] = doctor.id

            return redirect(url_for("doctor_dashboard"))

        return "Invalid Email or Password"

    return render_template("doctor_login.html")

@app.route("/doctor_dashboard")
def doctor_dashboard():

    if "doctor" not in session:
        return redirect(url_for("doctor_login"))

    doctor = Doctor.query.get(session["doctor"])

    return render_template(
        "doctor_dashboard.html",
        doctor=doctor
    )

@app.route("/register_conference/<int:id>")
def register_conference(id):

    if "doctor" not in session:
        return redirect(url_for("doctor_login"))

    existing = ConferenceRegistration.query.filter_by(
        doctor_id=session["doctor"],
        conference_id=id
    ).first()

    if existing:
        return redirect(url_for("my_conferences"))

    registration = ConferenceRegistration(
        doctor_id=session["doctor"],
        conference_id=id
    )

    db.session.add(registration)
    db.session.commit()

    return redirect(url_for("my_conferences"))

@app.route("/my_conferences")
def my_conferences():

    if "doctor" not in session:
        return redirect(url_for("doctor_login"))

    registrations = ConferenceRegistration.query.filter_by(
        doctor_id=session["doctor"]
    ).all()

    conference_list = []

    for registration in registrations:

        conference = Conference.query.get(
            registration.conference_id
        )

        conference_list.append(conference)

    return render_template(
        "my_conferences.html",
        conferences=conference_list
    )

if __name__ == "__main__":

    with app.app_context():
        db.create_all()

    app.run(debug=True)